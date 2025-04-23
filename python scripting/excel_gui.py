import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
import os

filename = "data_table.xlsx"

# এক্সেল ফাইল সেটআপ
def setup_excel():
    if not os.path.exists(filename):
        wb = openpyxl.Workbook()

        sheet1 = wb.active
        sheet1.title = "Chapter"
        sheet1.append(["ID", "Name"])

        sheet2 = wb.create_sheet("Section")
        sheet2.append(["ID", "Name"])

        sheet3 = wb.create_sheet("Hadis")
        sheet3.append(["ID", "Name"])

        wb.save(filename)

# অটো ID
def get_next_id(sheet):
    max_id = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None and isinstance(row[0], int):
            max_id = max(max_id, row[0])
    return max_id + 1

# ডেটা ইনসার্ট
def insert_data(sheet_name, name):
    try:
        wb = openpyxl.load_workbook(filename)
        if sheet_name not in wb.sheetnames:
            messagebox.showerror("Error", f"❌ Sheet '{sheet_name}' নেই!")
            return

        sheet = wb[sheet_name]
        next_id = get_next_id(sheet)
        sheet.append([next_id, name])
        wb.save(filename)
        messagebox.showinfo("Success", f"✅ ডেটা ইনসার্ট হয়েছে (ID: {next_id}) → Sheet: {sheet_name}")
        name_entry.delete(0, tk.END)

    except Exception as e:
        messagebox.showerror("Error", f"❌ ত্রুটি: {str(e)}")

# সাবমিট বাটনের ফাংশন
def submit():
    sheet_name = sheet_var.get()
    name = name_entry.get().strip()
    if not name:
        messagebox.showwarning("Warning", "👤 Name ইনপুট দিন!")
        return
    insert_data(sheet_name, name)

# GUI শুরু
setup_excel()
root = tk.Tk()
root.title("📘 Excel Data Inserter")
root.geometry("350x200")
root.resizable(False, False)

# শীট সিলেকশন
sheet_var = tk.StringVar(value="Chapter")
tk.Label(root, text="📄 Sheet নির্বাচন করুন:").pack(pady=5)
sheet_menu = ttk.Combobox(root, textvariable=sheet_var, values=["Chapter", "Section", "Hadis"], state="readonly")
sheet_menu.pack()

# নাম ইনপুট
tk.Label(root, text="👤 Name দিন:").pack(pady=5)
name_entry = tk.Entry(root, width=30)
name_entry.pack()

# সাবমিট বাটন
tk.Button(root, text="✅ Insert", command=submit, bg="#4CAF50", fg="white", width=20).pack(pady=15)

# ✅ Enter চাপলে সাবমিট হবে
root.bind('<Return>', lambda event: submit())

root.mainloop()
