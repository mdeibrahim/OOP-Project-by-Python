import openpyxl
import os

filename = "data_table.xlsx"

# এক্সেল ফাইল ও শীট তৈরি (না থাকলে)
def setup_excel():
    if not os.path.exists(filename):
        wb = openpyxl.Workbook()

        # Student Info শীট
        sheet1 = wb.active
        sheet1.title = "Chapter"
        sheet1.append(["ID", "Name"])

        # Section শীট
        sheet2 = wb.create_sheet("Section")
        sheet2.append(["ID", "Name"])

        #Hadis শীট
        sheet3=wb.create_sheet("Hadis")
        sheet3.append(["ID","Name"])

        wb.save(filename)

#  অটো ID তৈরি
def get_next_id(sheet):
    max_id = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None and isinstance(row[0], int):
            max_id = max(max_id, row[0])
    return max_id + 1

#  ডেটা ইনসার্ট ফাংশন
def insert_data(sheet_name, data_list):
    wb = openpyxl.load_workbook(filename)
    if sheet_name not in wb.sheetnames:
        print(f"❌ Sheet '{sheet_name}' নেই!")
        return

    sheet = wb[sheet_name]
    next_id = get_next_id(sheet)
    sheet.append([next_id] + data_list)
    wb.save(filename)
    print(f"✅ ডেটা ইনসার্ট হয়েছে (ID: {next_id}) → Sheet: {sheet_name}")

# মেইন প্রোগ্রাম
def main():
    setup_excel()
    print(" Sheet Available: Chapter / Section / Hadis")
    sheet_name = input(" কোন শীটে ডেটা ইনসার্ট করবেন? ").strip()

    if sheet_name == "Chapter":
        id = input("ID: ")
        name = input("Name: ")
        insert_data(sheet_name, [id, name])

    elif sheet_name == "Section":
        id = input("ID: ")
        name = input("Name: ")
        insert_data(sheet_name, [id, name])
    
    elif sheet_name == "Hadis":
        id = input("ID: ")
        name = input("Name: ")
        insert_data(sheet_name, [id, name])

    else:
        print(" শীটের নাম ভুল দিয়েছেন!")

if __name__ == "__main__":
    main()
